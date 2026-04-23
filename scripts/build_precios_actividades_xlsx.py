#!/usr/bin/env python3
"""Genera plantilla Excel de precios (sin IVA). Ejecutar desde la raíz del repo:
   .venv-xlsx/bin/python scripts/build_precios_actividades_xlsx.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "plantillas" / "precios-actividades-sin-iva.xlsx"

ROWS = [
    ("Tour Guiado Español", 0, 7, 0.15, 0.07),
    ("Tour Guiado Inglés", 0, 12, 0.15, 0.07),
    ("Tour Guiado Portugués", 0, 12, 0.15, 0.07),
    ("Tour Guiado Bici", 0, 12, 0.15, 0.07),
    ("Jack y Lupa", 5, None, 0.15, 0.07),
    ("Refugio + Toro + Traslado", 155, None, 0.15, 0.07),
    ("Bruma", 120, None, 0.10, 0.07),
    ("Misión + SIO", 40, None, 0.15, 0.07),
    ("Asado + Bote p/2", 220, None, 0.15, 0.07),
    ("Asado + Bote p/1", 110, None, 0.15, 0.07),
    ("Ejemplo costo alto (editar nombre)", 850, None, 0.15, 0.07),
]


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Hoja Notas ---
    ws0 = wb.active
    ws0.title = "Notas"
    ws0["A1"] = "Plantilla de precios — actividades turísticas exoneradas de IVA"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = (
        "Modelo: Precio sugerido = (Costo fijo + Costo guía) × (1 + Margen) ÷ (1 − Comisión Handy). "
        "Si no hay guía, dejá Costo guía vacío o 0."
    )
    ws0["A5"] = "Ganancia estimada = Precio × (1 − Comisión) − (Costo fijo + Costo guía)."
    ws0["A7"] = "Ajustá porcentajes como decimales (15% = 0,15) o usá formato Porcentaje en Excel."
    ws0.column_dimensions["A"].width = 100
    for r in range(1, 10):
        ws0[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Hoja Precios ---
    ws = wb.create_sheet("Precios", 0)
    headers = (
        "Actividad",
        "Costo fijo (USD)",
        "Costo guía (USD)",
        "Margen deseado",
        "Comisión Handy",
        "Precio final sugerido (USD)",
        "Ganancia estimada (USD)",
    )
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="ED8169")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (name, fixed, guide, margin, handy) in enumerate(ROWS, start=2):
        ws.cell(row=i, column=1, value=name).border = border
        ws.cell(row=i, column=2, value=fixed).border = border
        gcell = ws.cell(row=i, column=3, value=0 if guide is None else guide)
        gcell.border = border
        ws.cell(row=i, column=4, value=margin).border = border
        ws.cell(row=i, column=5, value=handy).border = border

        ws.cell(row=i, column=4).number_format = "0.0%"
        ws.cell(row=i, column=5).number_format = "0.0%"
        ws.cell(row=i, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=3).number_format = '"$"#,##0.00'

        # Fórmulas (notación en inglés, compatible con .xlsx)
        ws.cell(
            row=i,
            column=6,
            value=f"=(B{i}+C{i})*(1+D{i})/(1-E{i})",
        ).border = border
        ws.cell(row=i, column=6).number_format = '"$"#,##0.00'

        ws.cell(
            row=i,
            column=7,
            value=f"=F{i}*(1-E{i})-(B{i}+C{i})",
        ).border = border
        ws.cell(row=i, column=7).number_format = '"$"#,##0.00'

    widths = (34, 16, 16, 16, 18, 22, 22)
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"Escrito: {OUT}")


if __name__ == "__main__":
    main()
